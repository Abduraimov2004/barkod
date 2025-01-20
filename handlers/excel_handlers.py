# handlers/excel_handlers.py

import os
import openpyxl
from telegram import Update, File
from telegram.ext import ContextTypes
from states import States
from keyboards import import_export_keyboard, admin_menu_keyboard
from config import ADMIN_ID
from database import (
    add_product,
    get_all_products,
    get_product_by_barcode,
    update_weight_price_finalprice
)
import logging

logger = logging.getLogger(__name__)

XLSX_FILE = "exported_products.xlsx"

async def import_export_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("Siz admin emassiz!")
        return States.MAIN_MENU

    text = update.message.text
    if text == "XLSX import":
        await update.message.reply_text(
            "Iltimos, *XLSX faylni yuboring*.\n\n"
            "Ustunlar tartibi quyidagicha bo'lishi kerak (16 ta):\n"
            "`id, name, artikul, barcode, category, postavshik, stock, cena_postavki,`\n"
            "`cena_prodaji, skidka, brend, srok, edinitsa_izmereniya, weight, price, final_price`\n\n"
            "Eslatma: weight, price, final_price bo'yicha agar barkod mavjud bo'lsa - o'sha ustunlar keyinchalik barkod orqali yangilanadi.\n"
            "Agar new bo'lsa, DB ga qo'shishda weight/price/final_price ham kiritib qo'yiladi (lekin biz odatda barkod panelida yangilaymiz)."
        )
        return States.IMPORT_EXPORT

    elif text == "XLSX export":
        products = get_all_products(limit=999999, offset=0)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Column headers
        headers = [
            "id", "name", "artikul", "barcode", "category", "postavshik",
            "stock", "cena_postavki", "cena_prodaji", "skidka",
            "brend", "srok", "edinitsa_izmereniya",
            "weight", "price", "final_price"
        ]
        ws.append(headers)

        for row in products:
            ws.append([
                row.get('id'),
                row.get('name'),
                row.get('artikul'),
                row.get('barcode'),
                row.get('category'),
                row.get('postavshik'),
                row.get('stock'),
                row.get('cena_postavki'),
                row.get('cena_prodaji'),
                row.get('skidka'),
                row.get('brend'),
                row.get('srok'),
                row.get('edinitsa_izmereniya'),
                row.get('weight'),
                row.get('price'),
                row.get('final_price')
            ])

        wb.save(XLSX_FILE)

        if os.path.exists(XLSX_FILE):
            with open(XLSX_FILE, "rb") as f:
                await update.message.reply_document(document=f, filename=XLSX_FILE)
            os.remove(XLSX_FILE)  # Remove the file after sending

        return States.IMPORT_EXPORT

    elif text == "Orqaga":
        await update.message.reply_text(
            "Admin menyusiga qaytildi",
            reply_markup=admin_menu_keyboard()
        )
        return States.ADMIN_MENU

    else:
        await update.message.reply_text("Noto'g'ri buyruq.")
        return States.IMPORT_EXPORT

async def excel_import_file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("Siz admin emassiz!")
        return States.MAIN_MENU

    # Check if message contains a document
    doc = update.message.document
    if not doc:
        await update.message.reply_text("Fayl kelmadi yoki bu document emas.")
        return States.IMPORT_EXPORT

    # Check file format
    if not doc.file_name.lower().endswith('.xlsx'):
        await update.message.reply_text("Iltimos, xlsx formatdagi fayl yuboring!")
        return States.IMPORT_EXPORT

    # Define the path to save the file
    file_path = f"import_{doc.file_name}"

    try:
        # Get the file
        logger.info("Getting file from Telegram API...")

        telegram_file: File = await context.bot.get_file(doc.file_id)
        if not telegram_file:
            logger.error("[!] Telegram API did not return the file.")
            await update.message.reply_text("Telegram API faylni qaytara olmadi.")
            return States.IMPORT_EXPORT

        # Download the file
        logger.info("Downloading file...")
        await telegram_file.download_to_drive(file_path)

        logger.info(f"File successfully downloaded: {file_path}")

        # Process the Excel file
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            logger.error(f"[!] Error opening Excel file: {e}")
            await update.message.reply_text("Excel faylni ochishda xatolik yuz berdi.")
            os.remove(file_path)
            return States.IMPORT_EXPORT

        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue  # Skip header

            if not row:
                continue

            # Ensure the row has 16 columns
            row_values = list(row) + [None] * (16 - len(row)) if len(row) < 16 else list(row[:16])

            (
                id_val,
                name,
                artikul,
                barcode_val,
                category,
                postavshik,
                stock_val,
                cena_postavki_val,
                cena_prodaji_val,
                skidka_val,
                brend,
                srok,
                edi,
                weight_val,
                price_val,
                final_price_val
            ) = row_values

            # Check barcode
            if not barcode_val or not str(barcode_val).isdigit():
                continue

            barcode_val = int(barcode_val)

            existing = get_product_by_barcode(barcode_val)
            if existing:
                # Update weight, price, final_price if present
                updates = {}
                if weight_val is not None and isinstance(weight_val, (int, float)):
                    updates['weight'] = float(weight_val)
                if price_val is not None and isinstance(price_val, (int, float)):
                    updates['price'] = float(price_val)
                if final_price_val is not None and isinstance(final_price_val, (int, float)):
                    updates['final_price'] = float(final_price_val)
                if updates:
                    update_weight_price_finalprice(
                        barcode=barcode_val,
                        weight=updates.get('weight'),
                        price=updates.get('price'),
                        final_price=updates.get('final_price')
                    )
                continue  # Skip adding existing products

            # Convert types
            def to_float(x):
                try:
                    return float(x)
                except:
                    return 0.0

            def to_int(x):
                try:
                    return int(x)
                except:
                    return 0

            stock_val = to_float(stock_val)
            cena_postavki_val = to_int(cena_postavki_val)
            cena_prodaji_val = to_int(cena_prodaji_val)
            skidka_val = to_float(skidka_val)

            # Add product to DB
            add_product(
                name=name or "",
                artikul=artikul or "",
                barcode=barcode_val,
                category=category or "",
                postavshik=postavshik or "",
                stock=stock_val,
                cena_postavki=cena_postavki_val,
                cena_prodaji=cena_prodaji_val,
                skidka=skidka_val,
                brend=brend or "",
                srok=srok or "",
                edinitsa_izmereniya=edi or ""
            )

        await update.message.reply_text(
            "XLSX import yakunlandi. Yangi mahsulotlar DB ga qo'shildi (agar barcode mavjud bo'lmasa)."
        )
        os.remove(file_path)

    except Exception as e:
        # Log the exception
        logger.error(f"[!] General error: {e}")
        await update.message.reply_text("Xatolik yuz berdi. Logdan tekshiring.")
        if os.path.exists(file_path):
            os.remove(file_path)

    return States.IMPORT_EXPORT
